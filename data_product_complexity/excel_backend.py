import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.cell.cell import Cell
from .data_product_complexity import (
    DataProductComplexityAssessment,
    Section,
    Question,
    QuestionType,
    Backend,
)
from openpyxl.worksheet.cell_range import CellRange
import re


def apply_font_to_range(wb, range_str, bold=False, italic=False):
    """
    Applies bold and/or italic font to all cells in the specified range.

    Parameters:
        wb        : openpyxl Workbook object
        range_str : string like "Sheet1!A5:B11"
        bold      : apply bold font if True
        italic    : apply italic font if True
    """
    # Parse sheet name and range
    if "!" not in range_str:
        raise ValueError("Range must be in 'Sheet!A1:B2' format")

    sheet_name, cell_range = range_str.split("!")
    ws = wb[sheet_name]

    # Get boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    # Apply font to each cell
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.font = Font(bold=bold, italic=italic)


def fit_col_width(worksheet, col: str) -> None:
    col_idx = column_index_from_string(col)
    max_length = 0

    for row in worksheet.iter_rows(
        min_row=1, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx
    ):
        for cell in row:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

    worksheet.column_dimensions[col].width = max_length + 2


class CellLocationHelper:
    q_to_options_range: dict[Question, str] = {}
    q_to_options_range_with_score: dict[Question, str] = {}
    q_to_questionnaire_cell: dict[Question, str] = {}

    def notify_question_data_pos(self, question: Question, start_cell: Cell):
        num_options = len(question.options)
        ws_name = start_cell.parent.title
        start_cell_for_options = f"{start_cell.column_letter}{start_cell.row+3}"
        end_cell_for_options = (
            f"'{start_cell.column_letter}{start_cell.row+3+num_options}"
        )
        end_call_for_scores = (
            f"'{get_column_letter(start_cell.column+1)}{start_cell.row+3+num_options}"
        )
        self.q_to_options_range[question.question_id] = (
            f"'{ws_name}'!{start_cell_for_options}:{end_cell_for_options}"
        )
        self.q_to_options_range_with_score[question.question_id] = (
            f"'{ws_name}'!{start_cell_for_options}:{end_call_for_scores}"
        )

    def notify_question_dropdown_pos(self, question: Question, cell: Cell):
        self.q_to_questionnaire_cell[question.question_id] = get_full_cell_address(cell)

    def get_dropdown_pos_for_question(self, question: Question) -> str:
        return self.q_to_questionnaire_cell[question.question_id]

    def get_options_and_scores_range_for_question(self, question: Question) -> str:
        return self.q_to_options_range_with_score[question.question_id]

    def get_options_range_for_question(self, question: Question) -> str:
        return self.q_to_options_range[question.question_id]


def get_full_cell_address(cell):
    """
    Get the 'SheetName'!$A$6 form of a cell address
    """
    return f"{cell.parent.title}!{cell.coordinate}"


def question_score_formula(options_cell_address, validation_formula) -> str:
    """
    validation_formula is the range given to the list DataValidation (a range of cells that form the valid options)

    Finds the index of the option (that will form the rank value of the answer) unless it's the
    last answer (always Not Sure). We have this so that if a question had options
    A, B, C, D, E, Not sure
    Then we return 2 if the answer is C, 0 if the answer is A, and 0.5 if the answer is Not sure (take the average val)
    """
    formula = f"=IF(MATCH({options_cell_address}, {validation_formula}, 0) - 1 = COUNTA({validation_formula})-1,0.5,(MATCH({options_cell_address}, {validation_formula}, 0) - 1)/(COUNTA({validation_formula})-1))"

    return formula


class DataSheetBuilder:

    _cell_location_helper: CellLocationHelper

    def __init__(self, cell_location_helper: CellLocationHelper):
        self._cell_location_helper = cell_location_helper

    @staticmethod
    def _sanitize_sheet_name(name):
        """Convert to lowercase and keep only _0-9a-zA-Z"""
        base = "_data_" + re.sub(r"[^0-9a-zA-Z_]", "", name.replace(" ", "_").lower())
        return base[:31]  # Excel sheet name limit

    def _populate_data_sheet(self, ws: Worksheet, section: Section, index: int):
        """
        For each of the categories we create a hidden _data_categoryname tab that has the questions and options laid
        out like:

        |             | Question_1 | Question_2 | .... | Question_n |
        | Title       | xxxx       | xxxx       | .... | ...        |
        | Description | ....
        | NumOptions  | ....
        | Option_1    | ...  |
        | Option_2    | ...  |
        | ...
        | Option_n.   | ...

        After this we set up the rest of the Excel file to calculate everything from the
        hidden data tabs (DataValidations for dropdowns, understanding how many options in each question,
        how many questions in a section etc.)
        """
        sheet_name = ws.title
        category = section.title
        questions = section.questions
        headers = [""] + [f"Question_{x}" for x in range(1, len(questions) + 1)]
        titles = ["Title"] + [q.question_text for q in questions]
        num_questions = len(questions)
        descriptions = ["Description"] + [q.description for q in questions]
        max_options = max(len(q.options) for q in questions)
        num_options_row = ["NumOptions"] + [
            len(q.options) for q in questions
        ]
        option_rows = []

        for i in range(max_options):
            row = [f"Option_{i}"]
            for q in questions:
                opts = q.options
                row.append(opts[i].option_text if i < len(opts) else "")
            option_rows.append(row)

        for idx, q in enumerate(section.questions):
            self._cell_location_helper.notify_question_data_pos(
                question=q, start_cell=ws.cell(row=2, column=idx + 1)
            )

        ws.append(headers)
        ws.append(titles)
        ws.append(descriptions)
        ws.append(num_options_row)
        for row in option_rows:
            ws.append(row)

        # apply_font_to_range(wb, f"{sheet_name}!A1:A{max_options+4}", bold=True)
        # apply_font_to_range(
        #     wb, f"{sheet_name}!A1:{get_column_letter(num_questions+1)}1", bold=True
        # )
        for col in [get_column_letter(x) for x in range(1, num_questions + 1)]:
            fit_col_width(ws, col)

    def build(self, wb: Workbook, product: DataProductComplexityAssessment):
        for section_index, section in enumerate(product.scorable_sections):
            ws_data = wb.create_sheet(
                title=DataSheetBuilder._sanitize_sheet_name(section.title)
            )
            ws_data.sheet_state = "hidden"
            self._populate_data_sheet(ws_data, section, section_index)


class ScoreSheetBuilder:
    _ws: Worksheet
    _cell_location_helper: CellLocationHelper

    def __init__(self, ws: Worksheet, cell_location_helper: CellLocationHelper):
        self._ws = ws
        self._cell_location_helper = cell_location_helper

    def _formula_for_question(self, question: Question) -> str:
        cell_addr = self._cell_location_helper.get_dropdown_pos_for_question(question)
        options_and_scores_range = (
            self._cell_location_helper.get_options_and_scores_range_for_question(
                question
            )
        )
        return f"=VLOOKUP({cell_addr}, {options_and_scores_range},2,FALSE)"

    def build(self, product: DataProductComplexityAssessment) -> None:
        row = 2  # Start from row 2 to leave space for headers

        # Headers
        ws_score.cell(row=1, column=1, value="Section")
        ws_score.cell(row=1, column=2, value="Final Score (1–5)")
        ws_score.cell(row=1, column=1).font = Font(bold=True)
        ws_score.cell(row=1, column=2).font = Font(bold=True)

        for category in question_formulas.keys():
            if category == "Data Product Information":
                pass
            else:
                ws_score.cell(row=row, column=1, value=category)
                for qidx, question in enumerate(question_formulas[category]):
                    tmp_cell = ws_score.cell(row=row, column=3 + qidx)
                    tmp_cell.value = question_formulas[category][qidx]
                    # print(question_formulas[category][qidx])
                avg_formula = f"=INT(AVERAGE(C{row}:{get_column_letter(len(question_formulas[category])+3)}{row}) * 5) + 1"
                ws_score.cell(row=row, column=2).value = avg_formula
                row += 1

        # Apply heatmap conditional formatting to column B
        heatmap = ColorScaleRule(
            start_type="num",
            start_value=1,
            start_color="92D050",
            mid_type="num",
            mid_value=3,
            mid_color="FFFF00",
            end_type="num",
            end_value=5,
            end_color="FF0000",
        )
        ws_score.conditional_formatting.add(f"B2:B{row - 1}", heatmap)

        fit_col_width(worksheet=ws_score, col="A")
        fit_col_width(worksheet=ws_score, col="B")

        for col_letter in [get_column_letter(x) for x in range(3, 30)]:
            ws_score.column_dimensions[col_letter].hidden = True

        chart = BarChart()
        chart.type = "bar"
        chart.style = 10  # Excel predefined chart style
        chart.title = "Data Product Complexity Scores"
        chart.y_axis.title = "Sections"
        chart.x_axis.title = "Score"
        chart.y_axis.majorGridlines = None  # Optional: remove vertical gridlines
        chart.x_axis.majorGridlines = (
            chart.x_axis.majorGridlines
        )  # Ensures horizontal gridlines are visible

        # Reference data
        data_ref = Reference(ws_score, min_col=2, min_row=1, max_row=8)
        categories_ref = Reference(ws_score, min_col=1, min_row=2, max_row=8)

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(categories_ref)

        # Layout to roughly center chart (H1 is a good anchor)
        chart.layout = Layout(
            manualLayout=ManualLayout(
                x=0.25,
                y=0.1,
                h=0.6,
                w=0.5,  # Adjust for position and size
                xMode="factor",
                yMode="factor",
                hMode="factor",
                wMode="factor",
            )
        )

        # === Step 3: Insert Chart into Sheet ===
        ws_score.add_chart(chart, "H1")  # Center-ish position


class QuestionnaireSheetBuilder:
    _ws: Worksheet
    _clh: CellLocationHelper

    col_indexes = {"q_num": 1, "title": 2, "description": 2, "options": 3}

    def __init__(self, ws: Worksheet, cell_location_helper: CellLocationHelper):
        self._ws = ws
        self._clh = cell_location_helper

    def _render_dropdown_question(
        self, cursor_row: int, question: Question, section_num: int, question_num: int
    ) -> int:
        """
        Return number of rows added
        """
        # assert question.question_type == QuestionType.DROPDOWN

        # Add question number
        q_num = f"{section_num}.{question_num}"
        self._ws.cell(row=cursor_row, column=self.col_indexes["q_num"], value=q_num)

        # Add question title
        self._ws.cell(
            row=cursor_row,
            column=self.col_indexes["title"],
            value=question.question_text,
        )
        # Add description in row below
        d_cell = self._ws.cell(
            row=cursor_row + 1,
            column=self.col_indexes["description"],
            value=question.description,
        )
        d_cell.font = Font(italic=True)

        # Add dropdown of options
        options = question.options
        if options:
            dv = DataValidation(
                type="list",
                formula1=f'"{self._clh.get_options_range_for_question(question=question)}"',
                showDropDown=False,
            )
            dropdown_cell = self._ws.cell(
                row=cursor_row, column=self.col_indexes["options"]
            )        
            self._clh.notify_question_dropdown_pos(question=question, cell=dropdown_cell)

            self._ws.add_data_validation(dv)
            dv.add(dropdown_cell)

            # Default to Not sure option
            if "Not sure" in options:
                dropdown_cell.value = "Not sure"
        
        
        return 2

    def _render_section_heading(
        self, section: Section, row: int, section_number: 1
    ) -> int:
        """
        Returns the number of rows added
        """
        self._ws.cell(
            row=row, column=self.col_indexes["title"], value=f"{section.title}"
        )
        self._ws.cell(row=row, column=self.col_indexes["title"]).font = Font(bold=True)
        q_num_cell = self._ws.cell(
            row=row, column=self.col_indexes["q_num"], value=str(section_number)
        )
        q_num_cell.font = Font(bold=True)
        return 1

    def _render_data_product_info_section(
        self, section: Section, cursor_row: int
    ) -> int:
        """
        In the Questionnaire worksheet, we populate the "Data Product Info" section
        right at the top

        return: the number of rows added
        """
        rows_added = self._render_section_heading(section, cursor_row, 1)
        return rows_added + 1  # Add a blank row after

    def _render_scorable_section(
        self, section: Section, cursor_row: int, section_number: int
    ) -> int:
        """
        Populate a "scorable" section

        return: the number of rows added
        """
        rows_added = self._render_section_heading(section, cursor_row, section_number)
        for question_num, question in enumerate(section.questions, start=1):
            rows_added += self._render_dropdown_question(
                question=question,
                section_num=section_number,
                question_num=question_num,
                cursor_row=cursor_row + rows_added,
            )
        return rows_added + 1  # Add a blank row after

    def build(self, product: DataProductComplexityAssessment) -> None:
        """
        Populates the questions data from the questionnaire yaml,
        setting up the DataValidation (dropdowns) of options for each
        question as

        1   | Section title |
        1.1 | Question 1 title | Cell with Options in dropdown
            | Question 1 description
        """
        cursor_row = 1
        cursor_row += self._render_data_product_info_section(
            product.data_product_info, cursor_row
        )

        for section_index, section in enumerate(product.scorable_sections, start=1):
            cursor_row += self._render_scorable_section(
                section=section, cursor_row=cursor_row, section_number=section_index + 1
            )

        not_sure_fill = PatternFill(
            start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"
        )
        question_answer_range = f"B1:B{self._ws.max_row}"
        self._ws.conditional_formatting.add(
            question_answer_range,
            CellIsRule(operator="equal", formula=['"Not sure"'], fill=not_sure_fill),
        )

        fit_col_width(worksheet=self._ws, col="A")
        fit_col_width(worksheet=self._ws, col="B")
        fit_col_width(worksheet=self._ws, col="C")


class ExcelBackend(Backend):
    @staticmethod
    def _insert_new_sheet_at_pos(wb: Workbook, sheet_name: str, pos=1) -> Worksheet:
        ws = wb.create_sheet(sheet_name)
        wb._sheets.remove(ws)
        wb._sheets.insert(pos, ws)
        return ws

    @staticmethod
    def _strip_ampersands_from_formulae(wb: Workbook) -> None:
        # Need to strip @s from pyopenxsl's pesky formula validation
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        cell.value = cell.value.replace("@", "")

    def render(self, data: DataProductComplexityAssessment, output_path: str):
        wb = Workbook()
        clh = CellLocationHelper()
        data_sheets = DataSheetBuilder(clh)
        data_sheets.build(wb, data)

        question_sheet_builder = QuestionnaireSheetBuilder(
            ExcelBackend._insert_new_sheet_at_pos(wb, "Questions", 1), clh
        )
        question_sheet_builder.build(data)

        score_sheet_builder = ScoreSheetBuilder(
            ExcelBackend._insert_new_sheet_at_pos(wb, "Score", 2), clh
        )
        score_sheet_builder.build(data)

        del wb["Sheet"]
        wb.save("tmp_" + output_path)

        wb = load_workbook("tmp_" + output_path, data_only=False)
        ExcelBackend._strip_ampersands_from_formulae(wb)
        wb.save(output_path)
